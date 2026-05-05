import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from ocr_parser_netstal import LEFT_TEMPLATE, RIGHT_TEMPLATE

_FORMULA_CHARS = ("=", "+", "-", "@", "|", "\t", "\r")

def _safe_cell(value) -> str:
    s = str(value) if value is not None else ""
    if s and s[0] in _FORMULA_CHARS:
        return "'" + s
    return s

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left():
    return Alignment(horizontal="left", vertical="center")

HEADER_FILL = PatternFill("solid", fgColor="1A3C6E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
GROUP_FILL  = PatternFill("solid", fgColor="D9E4F0")
GROUP_FONT  = Font(bold=True, size=9)
DATA_FONT   = Font(size=9)
TITLE_FONT  = Font(bold=True, size=11)
META_FONT   = Font(size=9)

# Layout: A(項目) B(位置) C-D(名稱) E-F(設定值) G(範圍) H(單位)
#         I(項目) J(位置) K-L(名稱) M-N(設定值) O(範圍) P(單位)
LEFT_COLS  = list("ABCDEFGH")
RIGHT_COLS = list("IJKLMNOP")
RIGHT_GROUPS = {"機器加熱設定", "模具加熱設定", "乾燥機", "料桶"}


def create_blank_template() -> str:
    params = [{"group": g, "pos": p, "name": n, "value": "", "range": r, "unit": u}
              for g, p, n, r, u in LEFT_TEMPLATE + RIGHT_TEMPLATE]
    return _build_excel(
        {"機型": "", "日期": "", "克重": "", "瓶口": "",
         "模號": "", "穴數": "", "原料": "", "params": params},
        blank=True,
    )


def fill_template(data: dict) -> str:
    return _build_excel(data, blank=False)


def _checkbox_title(machine_type: str, blank: bool) -> str:
    if blank:
        return "Netstal 製程管制標準     □ EM09  □ EM13"
    marks = {opt: ("■" if opt == machine_type else "□") for opt in ["EM09", "EM13"]}
    return f"Netstal 製程管制標準     {marks['EM09']} EM09  {marks['EM13']} EM13"


def _build_excel(data: dict, blank: bool = False) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "製程管制標準"

    col_widths = {"A": 8, "B": 7, "C": 8, "D": 10, "E": 7, "F": 4,
                  "G": 7, "H": 7, "I": 8, "J": 7, "K": 8, "L": 10,
                  "M": 7, "N": 4, "O": 7, "P": 7}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: Title
    ws.merge_cells("A1:P1")
    ws["A1"] = _checkbox_title(data.get("機型", ""), blank)
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 18

    # Row 2: Meta
    ws.merge_cells("A2:D2")
    ws.merge_cells("E2:H2")
    ws.merge_cells("I2:L2")
    ws.merge_cells("M2:P2")
    穴數 = data.get("穴數", "")
    穴數_str = f"  {穴數}穴" if 穴數 else ""
    ws["A2"] = f'克重：{data.get("克重", "")}  瓶口：{data.get("瓶口", "")}'
    ws["E2"] = f'模號 {data.get("模號", "")}{穴數_str}'
    ws["I2"] = f'原料：{data.get("原料", "")}'
    ws["M2"] = _safe_cell(data.get("日期", ""))
    for cell in ["A2", "E2", "I2", "M2"]:
        ws[cell].font = META_FONT
        ws[cell].alignment = _left()
    ws.row_dimensions[2].height = 14

    # Row 3: Headers
    for col, label in zip(LEFT_COLS, ["項目", "位置", "名稱", None, "設定值", None, "範圍", "單位"]):
        if label:
            c = ws[f"{col}3"]
            c.value = label; c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = _center(); c.border = _thin()
    ws.merge_cells("C3:D3")
    ws.merge_cells("E3:F3")

    for col, label in zip(RIGHT_COLS, ["項目", "位置", "名稱", None, "設定值", None, "範圍", "單位"]):
        if label:
            c = ws[f"{col}3"]
            c.value = label; c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = _center(); c.border = _thin()
    ws.merge_cells("K3:L3")
    ws.merge_cells("M3:N3")
    ws.row_dimensions[3].height = 14

    params       = data.get("params", [])
    left_params  = [p for p in params if p.get("group") not in RIGHT_GROUPS]
    right_params = [p for p in params if p.get("group")     in RIGHT_GROUPS]
    max_rows     = max(len(left_params), len(right_params))

    for idx in range(max_rows):
        er = 4 + idx
        ws.row_dimensions[er].height = 13
        if idx < len(left_params):
            _write_left_row(ws, er, left_params[idx])
        else:
            ws.merge_cells(f"C{er}:D{er}")
            ws.merge_cells(f"E{er}:F{er}")
            for c in LEFT_COLS: ws[f"{c}{er}"].border = _thin()
        if idx < len(right_params):
            _write_right_row(ws, er, right_params[idx])
        else:
            ws.merge_cells(f"K{er}:L{er}")
            ws.merge_cells(f"M{er}:N{er}")
            for c in RIGHT_COLS: ws[f"{c}{er}"].border = _thin()

    _merge_groups(ws, left_params,  4, "A")
    _merge_groups(ws, right_params, 4, "I")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


def _write_left_row(ws, row, param):
    ws.merge_cells(f"C{row}:D{row}")
    ws.merge_cells(f"E{row}:F{row}")
    for col in LEFT_COLS:
        ws[f"{col}{row}"].border = _thin()
        ws[f"{col}{row}"].font   = DATA_FONT
    ws[f"A{row}"].value     = param["group"]
    ws[f"A{row}"].fill      = GROUP_FILL
    ws[f"A{row}"].font      = GROUP_FONT
    ws[f"A{row}"].alignment = _center(wrap=True)
    ws[f"B{row}"].value     = _safe_cell(param.get("pos", "")); ws[f"B{row}"].alignment = _center()
    ws[f"C{row}"].value     = _safe_cell(param["name"]);         ws[f"C{row}"].alignment = _left()
    ws[f"E{row}"].value     = _safe_cell(param["value"]);        ws[f"E{row}"].alignment = _center()
    ws[f"G{row}"].value     = _safe_cell(param["range"]);        ws[f"G{row}"].alignment = _center()
    ws[f"H{row}"].value     = _safe_cell(param["unit"]);         ws[f"H{row}"].alignment = _center()


def _write_right_row(ws, row, param):
    ws.merge_cells(f"K{row}:L{row}")
    ws.merge_cells(f"M{row}:N{row}")
    for col in RIGHT_COLS:
        ws[f"{col}{row}"].border = _thin()
        ws[f"{col}{row}"].font   = DATA_FONT
    ws[f"I{row}"].value     = param["group"]
    ws[f"I{row}"].fill      = GROUP_FILL
    ws[f"I{row}"].font      = GROUP_FONT
    ws[f"I{row}"].alignment = _center(wrap=True)
    ws[f"J{row}"].value     = _safe_cell(param.get("pos", "")); ws[f"J{row}"].alignment = _center()
    ws[f"K{row}"].value     = _safe_cell(param["name"]);         ws[f"K{row}"].alignment = _left()
    ws[f"M{row}"].value     = _safe_cell(param["value"]);        ws[f"M{row}"].alignment = _center()
    ws[f"O{row}"].value     = _safe_cell(param["range"]);        ws[f"O{row}"].alignment = _center()
    ws[f"P{row}"].value     = _safe_cell(param["unit"]);         ws[f"P{row}"].alignment = _center()


def _merge_groups(ws, params, start_row, col):
    if not params:
        return
    i = 0
    while i < len(params):
        group = params[i]["group"]
        j = i + 1
        while j < len(params) and params[j]["group"] == group:
            j += 1
        r1, r2 = start_row + i, start_row + j - 1
        if r2 > r1:
            ws.merge_cells(f"{col}{r1}:{col}{r2}")
        c = ws[f"{col}{r1}"]
        c.value = group; c.fill = GROUP_FILL; c.font = GROUP_FONT
        c.alignment = _center(wrap=True); c.border = _thin()
        i = j
