import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("日期",           6),
    ("機台編號",       9),
    ("工單號碼",      10),
    ("品名、蓋型",    16),
    ("產量(個)",      10),
    ("不良數(個)",    10),
    ("不良率(%)",     10),
    ("有無添加回收料", 13),
    ("廢蓋(KG)",       9),
    ("廢料(KG)",       9),
    ("可回收廢蓋(KG)", 13),
    ("不良原因",      14),
]

DATA_ROWS = 20

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def create_blank_template(process="切割"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "不良率紀錄"

    n_cols = len(COLUMNS)
    last_col = get_column_letter(n_cols)

    # Row 1: title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "塑蓋廠各站別不良率紀錄表"
    c.font = Font(size=16, bold=True)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Row 2: meta
    ws["A2"] = process
    ws["A2"].font = Font(bold=True)
    ws["D2"] = "組別："
    ws["E2"] = ""
    ws["G2"] = "班別："
    ws["H2"] = ""
    ws["J2"] = "年份："
    ws["K2"] = ""
    ws.row_dimensions[2].height = 20

    # Row 3: headers
    for i, (col_name, col_width) in enumerate(COLUMNS):
        col_letter = get_column_letter(i + 1)
        cell = ws.cell(row=3, column=i + 1, value=col_name)
        cell.font = Font(bold=True, size=10)
        cell.alignment = _center(wrap=True)
        cell.border = _thin()
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        ws.column_dimensions[col_letter].width = col_width
    ws.row_dimensions[3].height = 36

    # Data rows
    for row in range(4, 4 + DATA_ROWS):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin()
            cell.alignment = _center()
        ws.row_dimensions[row].height = 18

    # Footer
    footer_row = 4 + DATA_ROWS
    ws.merge_cells(f"A{footer_row}:C{footer_row}")
    ws[f"A{footer_row}"] = "廠長："
    ws.merge_cells(f"E{footer_row}:G{footer_row}")
    ws[f"E{footer_row}"] = "副廠："
    ws.merge_cells(f"I{footer_row}:{last_col}{footer_row}")
    ws[f"I{footer_row}"] = "填表人："
    ws.row_dimensions[footer_row].height = 20

    path = "/tmp/塑蓋廠不良率紀錄表_blank.xlsx"
    wb.save(path)
    return path


def fill_template(data):
    process = data.get("製程", "切割")
    path = create_blank_template(process)
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    ws["E2"] = data.get("組別", "")
    ws["H2"] = data.get("班別", "")
    ws["K2"] = data.get("年份", "")

    col_keys = [
        "日期", "機台編號", "工單號碼", "品名蓋型",
        "產量", "不良數", "不良率", "有無添加回收料",
        "廢蓋KG", "廢料KG", "可回收廢蓋KG", "不良原因",
    ]

    for i, row in enumerate(data.get("rows", [])):
        excel_row = 4 + i
        if excel_row >= 4 + DATA_ROWS:
            break
        for j, key in enumerate(col_keys):
            ws.cell(row=excel_row, column=j + 1).value = row.get(key, "")

    footer_row = 4 + DATA_ROWS
    ws[f"A{footer_row}"] = f'廠長：{data.get("廠長", "")}'
    ws[f"E{footer_row}"] = f'副廠：{data.get("副廠", "")}'
    ws[f"I{footer_row}"] = f'填表人：{data.get("填表人", "")}'

    out = "/tmp/不良率紀錄表_filled.xlsx"
    wb.save(out)
    return out
