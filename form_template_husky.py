import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from ocr_parser_husky import LEFT_TEMPLATE, RIGHT_TEMPLATE


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left():
    return Alignment(horizontal="left", vertical="center")

HEADER_FILL = PatternFill("solid", fgColor="1A3C6E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
GROUP_FILL  = PatternFill("solid", fgColor="D9E4F0")
GROUP_FONT  = Font(bold=True, size=9)
DATA_FONT   = Font(size=9)
TITLE_FONT  = Font(bold=True, size=11)
META_FONT   = Font(size=9)

# Layout: A(項目) B-C(名稱) D-E(設定值) F(範圍) G(單位) | H(項目) I-J(名稱) K-L(設定值) M(範圍) N(單位)
LEFT_COLS  = list("ABCDEFG")
RIGHT_COLS = list("HIJKLMN")
RIGHT_GROUPS = {"機器加熱設定", "模具加熱設定", "控制", "乾燥機"}


def create_blank_template(machine_type="EM04") -> str:
    params = [{"group": g, "name": n, "value": "", "range": r, "unit": u}
              for g, n, r, u in LEFT_TEMPLATE + RIGHT_TEMPLATE]
    return _build_excel({
        "機型": machine_type, "日期": "", "克重": "", "瓶口": "",
        "模號": "", "穴數": "", "原料": "",
        "params": params,
    })


def fill_template(data: dict) -> str:
    return _build_excel(data)


def _checkbox_title(machine_type: str) -> str:
    marks = {opt: ("■" if opt == machine_type else "□") for opt in ["EM04", "EM06", "EM07"]}
    return (f"Husky 製程管制標準     "
            f"{marks['EM04']} EM04  {marks['EM06']} EM06  {marks['EM07']} EM07")


def _build_excel(data: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "製程管制標準"

    col_widths = {"A": 9, "B": 6, "C": 14, "D": 8, "E": 4,
                  "F": 7, "G": 7, "H": 9, "I": 6, "J": 14,
                  "K": 8, "L": 4, "M": 7, "N": 7, "O": 2, "P": 2}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row 1: Title with EM04/EM06/EM07 checkboxes
    ws.merge_cells("A1:N1")
    ws["A1"] = _checkbox_title(data.get("機型", "EM04"))
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 18

    # Row 2: Meta fields
    ws.merge_cells("A2:D2")
    ws.merge_cells("E2:H2")
    ws.merge_cells("I2:L2")
    ws.merge_cells("M2:N2")
    穴數 = data.get("穴數", "")
    穴數_str = f"  {穴數}穴" if 穴數 else ""
    ws["A2"] = f'克重：{data.get("克重", "")}  瓶口：{data.get("瓶口", "")}'
    ws["E2"] = f'模號 {data.get("模號", "")}{穴數_str}'
    ws["I2"] = f'原料：{data.get("原料", "")}'
    ws["M2"] = data.get("日期", "")
    for cell in ["A2", "E2", "I2", "M2"]:
        ws[cell].font = META_FONT
        ws[cell].alignment = _left()
    ws.row_dimensions[2].height = 14

    # Row 3: Column headers
    for col, label in zip(LEFT_COLS, ["項目", "名稱", None, "設定值", None, "範圍", "單位"]):
        if label:
            c = ws[f"{col}3"]
            c.value = label; c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = _center(); c.border = _thin()
    ws.merge_cells("B3:C3")
    ws.merge_cells("D3:E3")

    for col, label in zip(RIGHT_COLS, ["項目", "名稱", None, "設定值", None, "範圍", "單位"]):
        if label:
            c = ws[f"{col}3"]
            c.value = label; c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = _center(); c.border = _thin()
    ws.merge_cells("I3:J3")
    ws.merge_cells("K3:L3")
    ws.row_dimensions[3].height = 14

    params = data.get("params", [])
    left_params  = [p for p in params if p["group"] not in RIGHT_GROUPS]
    right_params = [p for p in params if p["group"]     in RIGHT_GROUPS]

    # Separate out last left param (分配器打開延遲) — it spans A:C
    if left_params and left_params[-1]["name"] == "分配器打開延遲":
        regular_left = left_params[:-1]
        last_left    = left_params[-1]
    else:
        regular_left = left_params
        last_left    = None

    max_rows = max(len(regular_left), len(right_params))

    for idx in range(max_rows):
        er = 4 + idx
        ws.row_dimensions[er].height = 13
        if idx < len(regular_left):
            _write_left_row(ws, er, regular_left[idx])
        if idx < len(right_params):
            _write_right_row(ws, er, right_params[idx])

    # Special last row: 分配器打開延遲 (name spans A:C)
    if last_left:
        er = 4 + max_rows
        ws.row_dimensions[er].height = 13
        ws.merge_cells(f"A{er}:C{er}")
        ws.merge_cells(f"D{er}:E{er}")
        for col in LEFT_COLS:
            ws[f"{col}{er}"].border = _thin()
            ws[f"{col}{er}"].font   = DATA_FONT
        ws[f"A{er}"].value     = last_left["name"]
        ws[f"A{er}"].fill      = GROUP_FILL
        ws[f"A{er}"].font      = GROUP_FONT
        ws[f"A{er}"].alignment = _center()
        ws[f"D{er}"].value     = last_left["value"]
        ws[f"D{er}"].alignment = _center()
        ws[f"F{er}"].value     = last_left["range"]
        ws[f"F{er}"].alignment = _center()
        ws[f"G{er}"].value     = last_left["unit"]
        ws[f"G{er}"].alignment = _center()
        ws.merge_cells(f"H{er}:N{er}")
        ws[f"H{er}"].border = _thin()

    _merge_groups(ws, regular_left,  4, "A")
    _merge_groups(ws, right_params,  4, "H")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


def _write_left_row(ws, row, param):
    ws.merge_cells(f"B{row}:C{row}")
    ws.merge_cells(f"D{row}:E{row}")
    for col in LEFT_COLS:
        ws[f"{col}{row}"].border = _thin()
        ws[f"{col}{row}"].font   = DATA_FONT
    ws[f"A{row}"].value     = param["group"]
    ws[f"A{row}"].fill      = GROUP_FILL
    ws[f"A{row}"].font      = GROUP_FONT
    ws[f"A{row}"].alignment = _center(wrap=True)
    ws[f"B{row}"].value     = param["name"];  ws[f"B{row}"].alignment = _left()
    ws[f"D{row}"].value     = param["value"]; ws[f"D{row}"].alignment = _center()
    ws[f"F{row}"].value     = param["range"]; ws[f"F{row}"].alignment = _center()
    ws[f"G{row}"].value     = param["unit"];  ws[f"G{row}"].alignment = _center()


def _write_right_row(ws, row, param):
    ws.merge_cells(f"I{row}:J{row}")
    ws.merge_cells(f"K{row}:L{row}")
    for col in RIGHT_COLS:
        ws[f"{col}{row}"].border = _thin()
        ws[f"{col}{row}"].font   = DATA_FONT
    ws[f"H{row}"].value     = param["group"]
    ws[f"H{row}"].fill      = GROUP_FILL
    ws[f"H{row}"].font      = GROUP_FONT
    ws[f"H{row}"].alignment = _center(wrap=True)
    ws[f"I{row}"].value     = param["name"];  ws[f"I{row}"].alignment = _left()
    ws[f"K{row}"].value     = param["value"]; ws[f"K{row}"].alignment = _center()
    ws[f"M{row}"].value     = param["range"]; ws[f"M{row}"].alignment = _center()
    ws[f"N{row}"].value     = param["unit"];  ws[f"N{row}"].alignment = _center()


def _merge_groups(ws, params, start_row, col):
    if not params:
        return
    i = 0
    while i < len(params):
        group = params[i]["group"]
        j = i + 1
        while j < len(params) and params[j]["group"] == group:
            j += 1
        r1, r2 = start_row + i, start_row + j - 1
        if r2 > r1:
            ws.merge_cells(f"{col}{r1}:{col}{r2}")
        c = ws[f"{col}{r1}"]
        c.value = group; c.fill = GROUP_FILL; c.font = GROUP_FONT
        c.alignment = _center(wrap=True); c.border = _thin()
        i = j
